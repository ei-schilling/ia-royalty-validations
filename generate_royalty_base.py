#!/usr/bin/env python3
"""Generate ~500 realistic royalty settlement files for the knowledge base.

Formats: CSV (~200), XLSX (~150), JSON (~100), PDF-text-CSV (~50)
Each file simulates a royalty settlement statement with 10–80 rows.
Files deliberately include a mix of:
  - Clean / all-rules-pass files
  - Files with amount inconsistencies
  - Files with invalid rates, missing titles, bad dates
  - Files with duplicate entries, unknown transaction types
  - Files with guarantee/advance imbalances
  - Mixed-severity files
"""

from __future__ import annotations

import csv
import io
import json
import math
import os
import random
import string
from datetime import datetime, timedelta
from pathlib import Path

# Optional: openpyxl for xlsx
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ── Config ───────────────────────────────────────────────────────────────────

OUTPUT_DIR = Path(__file__).parent / "baseDocs" / "royaltyBase"
SEED = 42
random.seed(SEED)

TARGET_CSV = 200
TARGET_XLSX = 150
TARGET_JSON = 100
TARGET_PDF = 50  # pseudo-PDF (text files that mimic Schilling layout)

HEADERS = [
    "TRANSNR", "TRANSTYPE", "KONTO", "AFTALE", "ARTNR", "KANAL",
    "PRISGRUPPE", "VILKAR", "BILAGSNR", "BILAGSDATO", "ANTAL",
    "STKPRIS", "STKAFREGNSATS", "BELOEB", "VALUTA", "SKAT", "AFREGNBATCH",
]

# ── Reference data pools ────────────────────────────────────────────────────

VALID_TRANSTYPES = [
    "Salg", "Retur", "Frieksp", "Makulatur", "Svind", "Tilgang",
    "Overforsel", "Korrektion", "Efterreg", "Afgift", "Skat", "Moms",
    "Forskud", "ForskudMod", "GarGlobal", "GarGlobalMod", "GarLokal",
    "GarLokalMod", "Royalty", "RoyaltyMod", "Bonus", "BonusMod",
    "Afskrivning", "AfskrivningMod", "Udbetaling", "UdbetalingMod",
    "Indbetaling", "IndbetalingMod", "Overfort", "OverfortMod",
    "Rente", "RenteMod", "Diverse", "DiverseMod", "Rabat", "RabatMod",
]
DEPRECATED_TYPES = ["Speciel", "SpecielMod"]
INVALID_TYPES = ["XYZ123", "FooBar", "UnknownType", "MISC", "Placeholder"]

# Weighted: mostly Salg/Retur for realistic distribution
WEIGHTED_TYPES = (
    ["Salg"] * 40 + ["Retur"] * 10 + ["Frieksp"] * 3 + ["Korrektion"] * 3 +
    ["Royalty"] * 5 + ["Forskud"] * 3 + ["ForskudMod"] * 2 +
    ["GarGlobal"] * 2 + ["GarGlobalMod"] * 1 +
    ["Bonus"] * 2 + ["Afgift"] * 1 + ["Diverse"] * 1
)

AUTHORS = [
    ("AUTH-0042", "Hans Andersen"), ("AUTH-0108", "Karen Blixen"),
    ("AUTH-0215", "Peter Hoeg"), ("AUTH-0333", "Inger Christensen"),
    ("AUTH-0401", "Tove Ditlevsen"), ("AUTH-0519", "Henrik Pontoppidan"),
    ("AUTH-0627", "Isak Dinesen"), ("AUTH-0734", "Martin A. Hansen"),
    ("AUTH-0812", "Svend Aage Madsen"), ("AUTH-0999", "Naja Marie Aidt"),
    ("AUTH-1001", "Helle Helle"), ("AUTH-1102", "Josefine Klougart"),
    ("AUTH-1203", "Olga Ravn"), ("AUTH-1304", "Jonas Eika"),
    ("AUTH-1405", "Dorthe Nors"), ("AUTH-1506", "Ida Jessen"),
    ("AUTH-1607", "Jan Sonnergaard"), ("AUTH-1708", "Jens Blendstrup"),
    ("AUTH-1809", "Christina Hesselholdt"), ("AUTH-1910", "Lars Skinnebach"),
]

AGREEMENTS = [
    "AFT-2018-001", "AFT-2018-042", "AFT-2019-003", "AFT-2019-077",
    "AFT-2020-011", "AFT-2020-055", "AFT-2020-099", "AFT-2021-004",
    "AFT-2021-028", "AFT-2021-063", "AFT-2022-007", "AFT-2022-015",
    "AFT-2022-081", "AFT-2023-002", "AFT-2023-039", "AFT-2023-090",
    "AFT-2024-001", "AFT-2024-017", "AFT-2024-044", "AFT-2024-066",
    "AFT-2025-008", "AFT-2025-023", "AFT-2025-050", "AFT-2025-071",
]

KANALER = ["Boghandel", "Online", "Bibliotek", "Eksport", "Lydbog", "E-bog", "Subskription"]
PRISGRUPPER = ["Standard", "Pocket", "Indbundet", "Special", "Klassiker", "Lydbog", "Digital"]
VILKAAR = ["A", "B", "C", "D"]
CURRENCIES = ["DKK", "SEK", "NOK", "EUR"]
DATE_FORMATS = ["%d-%m-%Y", "%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"]

TITLES = [
    "Eventyr og Historier", "Den Afrikanske Farm", "Froken Smillas Fornemmelse for Sne",
    "Det", "Ansigterne", "Lykke-Per", "Vintereventyr", "Lykkelige Kristoffer",
    "Intet", "Gift", "Hest Hest Tiger Tiger", "De Usynlige",
    "Min Kamp Bind 1", "Babettes Gaestebud", "Kongens Fald",
    "Havbrevene", "Taler i Ode Kirker", "Den Stumme", "Mig Ejer Ingen",
    "Under Solen", "Drommen om det Rode Kammer", "Kunsten at Graede i Kor",
    "Nordkraft", "Tid og Rum", "At Fortaelle et Menneske", "Spejlinger",
    "Lysets Engel", "En By af Lys", "Krystaller", "Det Forsomte Forar",
]

def _gen_isbn() -> str:
    """Generate a valid ISBN-13 with 978-87 prefix."""
    prefix = [9, 7, 8, 8, 7]
    mid = [random.randint(0, 9) for _ in range(7)]
    digits = prefix + mid
    total = sum(d * (1 if i % 2 == 0 else 3) for i, d in enumerate(digits))
    check = (10 - total % 10) % 10
    all_d = digits + [check]
    s = "".join(map(str, all_d))
    return f"{s[:3]}-{s[3:5]}-{s[5:9]}-{s[9:12]}-{s[12]}"


def _gen_bad_isbn() -> str:
    """Generate an ISBN-13 with wrong check digit."""
    isbn = _gen_isbn()
    last = int(isbn[-1])
    bad = (last + random.randint(1, 9)) % 10
    return isbn[:-1] + str(bad)


def _random_date(start_year=2018, end_year=2026) -> datetime:
    start = datetime(start_year, 1, 1)
    end = datetime(end_year, 12, 31)
    delta = (end - start).days
    return start + timedelta(days=random.randint(0, delta))


def _format_date(dt: datetime, fmt: str | None = None) -> str:
    if fmt is None:
        fmt = random.choice(DATE_FORMATS)
    return dt.strftime(fmt)

# ── Scenario definitions ────────────────────────────────────────────────────

class Scenario:
    """Defines what kind of data errors a file should contain."""
    CLEAN = "clean"
    AMOUNT_MISMATCH = "amount_mismatch"
    INVALID_RATE = "invalid_rate"
    MISSING_TITLE = "missing_title"
    BAD_DATE = "bad_date"
    DUPLICATE = "duplicate"
    UNKNOWN_TYPE = "unknown_type"
    DEPRECATED_TYPE = "deprecated_type"
    HIGH_RATE = "high_rate"
    GUARANTEE_IMBALANCE = "guarantee_imbalance"
    ADVANCE_IMBALANCE = "advance_imbalance"
    NEGATIVE_PAYOUT = "negative_payout"
    BAD_ISBN = "bad_isbn"
    MIXED = "mixed"  # multiple issues

SCENARIO_WEIGHTS = {
    Scenario.CLEAN: 25,
    Scenario.AMOUNT_MISMATCH: 12,
    Scenario.INVALID_RATE: 8,
    Scenario.MISSING_TITLE: 8,
    Scenario.BAD_DATE: 7,
    Scenario.DUPLICATE: 8,
    Scenario.UNKNOWN_TYPE: 5,
    Scenario.DEPRECATED_TYPE: 4,
    Scenario.HIGH_RATE: 6,
    Scenario.GUARANTEE_IMBALANCE: 4,
    Scenario.ADVANCE_IMBALANCE: 4,
    Scenario.BAD_ISBN: 4,
    Scenario.MIXED: 15,
}


def _pick_scenario() -> str:
    population = list(SCENARIO_WEIGHTS.keys())
    weights = list(SCENARIO_WEIGHTS.values())
    return random.choices(population, weights)[0]


# ── Row generation ───────────────────────────────────────────────────────────

_transnr_counter = 10000


def _next_transnr() -> str:
    global _transnr_counter
    _transnr_counter += 1
    return str(_transnr_counter)


def _gen_row(
    scenario: str,
    author: tuple[str, str],
    agreement: str,
    base_date: datetime,
    batch: str,
    date_fmt: str,
    is_error_row: bool = False,
) -> dict:
    """Generate a single settlement row, optionally injecting errors."""

    transtype = random.choice(WEIGHTED_TYPES)
    kanal = random.choice(KANALER)
    prisgruppe = random.choice(PRISGRUPPER)
    vilkar = random.choice(VILKAAR)
    valuta = random.choice(CURRENCIES) if random.random() < 0.1 else "DKK"
    isbn = _gen_isbn()
    qty = random.randint(1, 2000)
    price = round(random.uniform(29.95, 599.00), 2)
    rate = round(random.uniform(0.05, 0.25), 4)
    amount = round(qty * price * rate, 2)
    skat = 0.00
    row_date = base_date + timedelta(days=random.randint(-30, 30))

    # Retur → negative quantities
    if transtype == "Retur":
        qty = -abs(qty)
        amount = round(qty * price * rate, 2)

    # Apply scenario-specific mutations
    if is_error_row:
        if scenario == Scenario.AMOUNT_MISMATCH:
            # Make beloeb wrong by 5-500
            amount = round(amount + random.uniform(5, 500) * random.choice([-1, 1]), 2)
        elif scenario == Scenario.INVALID_RATE:
            rate = random.choice([0, -0.05, None, "N/A"])
        elif scenario == Scenario.MISSING_TITLE:
            isbn = ""
        elif scenario == Scenario.BAD_DATE:
            row_date = datetime(random.choice([1899, 1950, 2199]), 1, 1)
        elif scenario == Scenario.DUPLICATE:
            pass  # duplicates handled at file level
        elif scenario == Scenario.UNKNOWN_TYPE:
            transtype = random.choice(INVALID_TYPES)
        elif scenario == Scenario.DEPRECATED_TYPE:
            transtype = random.choice(DEPRECATED_TYPES)
        elif scenario == Scenario.HIGH_RATE:
            rate = round(random.uniform(0.55, 0.95), 4)
            amount = round(qty * price * rate, 2)
        elif scenario == Scenario.BAD_ISBN:
            isbn = _gen_bad_isbn()
        elif scenario == Scenario.GUARANTEE_IMBALANCE:
            transtype = "GarGlobal"
            amount = round(random.uniform(1000, 50000), 2)
        elif scenario == Scenario.ADVANCE_IMBALANCE:
            transtype = "ForskudMod"
            amount = round(random.uniform(50000, 200000), 2)
        elif scenario == Scenario.NEGATIVE_PAYOUT:
            transtype = "Udbetaling"
            amount = -round(random.uniform(1000, 50000), 2)

    # Format rate for CSV
    rate_str = "" if rate is None else (str(rate) if isinstance(rate, (int, float)) else str(rate))

    bilagsnr = str(random.randint(50000, 99999))
    return {
        "TRANSNR": _next_transnr(),
        "TRANSTYPE": transtype,
        "KONTO": author[0],
        "AFTALE": agreement,
        "ARTNR": isbn,
        "KANAL": kanal,
        "PRISGRUPPE": prisgruppe,
        "VILKAR": vilkar,
        "BILAGSNR": bilagsnr,
        "BILAGSDATO": _format_date(row_date, date_fmt),
        "ANTAL": str(qty),
        "STKPRIS": f"{price:.2f}",
        "STKAFREGNSATS": rate_str,
        "BELOEB": f"{amount:.2f}",
        "VALUTA": valuta,
        "SKAT": f"{skat:.2f}",
        "AFREGNBATCH": batch,
    }


def _gen_file_rows(scenario: str, num_rows: int) -> list[dict]:
    """Generate a list of rows for one file."""
    author = random.choice(AUTHORS)
    agreement = random.choice(AGREEMENTS)
    base_date = _random_date()
    batch = str(random.randint(1000, 9999))
    date_fmt = random.choice(DATE_FORMATS)

    # For mixed: pick 2-4 sub-scenarios
    if scenario == Scenario.MIXED:
        sub_scenarios = random.sample(
            [s for s in SCENARIO_WEIGHTS if s not in (Scenario.CLEAN, Scenario.MIXED)],
            k=random.randint(2, 4),
        )
    else:
        sub_scenarios = [scenario]

    # Decide how many error rows (0 for clean, ~20-40% for others)
    if scenario == Scenario.CLEAN:
        error_indices = set()
    else:
        n_errors = max(1, int(num_rows * random.uniform(0.15, 0.4)))
        error_indices = set(random.sample(range(num_rows), min(n_errors, num_rows)))

    rows: list[dict] = []
    for i in range(num_rows):
        is_err = i in error_indices
        sc = random.choice(sub_scenarios) if is_err else Scenario.CLEAN
        row = _gen_row(sc, author, agreement, base_date, batch, date_fmt, is_error_row=is_err)
        rows.append(row)

    # Inject duplicates for DUPLICATE scenario
    if scenario == Scenario.DUPLICATE or (scenario == Scenario.MIXED and Scenario.DUPLICATE in sub_scenarios):
        n_dupes = random.randint(2, min(5, num_rows // 3))
        for _ in range(n_dupes):
            src = random.choice(rows[:max(1, len(rows) - 5)])
            dupe = dict(src)
            dupe["TRANSNR"] = _next_transnr()  # new transnr but same key fields
            rows.append(dupe)

    # For guarantee imbalance, add GarGlobal without matching GarGlobalMod
    if scenario == Scenario.GUARANTEE_IMBALANCE:
        rows.append(_gen_row(
            Scenario.CLEAN, author, agreement, base_date, batch, date_fmt
        ))
        rows[-1]["TRANSTYPE"] = "GarGlobal"
        rows[-1]["BELOEB"] = f"{round(random.uniform(5000, 30000), 2):.2f}"

    # For advance imbalance, add ForskudMod exceeding Forskud
    if scenario == Scenario.ADVANCE_IMBALANCE:
        # Add a small Forskud
        frow = _gen_row(Scenario.CLEAN, author, agreement, base_date, batch, date_fmt)
        frow["TRANSTYPE"] = "Forskud"
        frow["BELOEB"] = f"{round(random.uniform(1000, 5000), 2):.2f}"
        rows.append(frow)
        # Add a large ForskudMod
        mrow = _gen_row(Scenario.CLEAN, author, agreement, base_date, batch, date_fmt)
        mrow["TRANSTYPE"] = "ForskudMod"
        mrow["BELOEB"] = f"{round(random.uniform(10000, 50000), 2):.2f}"
        rows.append(mrow)

    return rows


# ── File writers ─────────────────────────────────────────────────────────────

def _write_csv(rows: list[dict], path: Path, delimiter: str = ";"):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=HEADERS, delimiter=delimiter, quoting=csv.QUOTE_ALL)
        writer.writeheader()
        writer.writerows(rows)


def _write_json(rows: list[dict], path: Path):
    # Randomly choose between flat array and {rows: [...]} format
    if random.random() < 0.4:
        data = rows
    else:
        data = {"rows": rows}
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def _write_xlsx(rows: list[dict], path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Royalty Settlement"
    ws.append(HEADERS)
    for row in rows:
        ws.append([row[h] for h in HEADERS])
    wb.save(path)


def _write_pseudo_pdf(rows: list[dict], path: Path):
    """Write a Schilling-style royalty settlement as a real PDF.

    Uses fpdf2 to generate a valid PDF with the same structured layout
    that Schilling royalty PDFs use — suitable for RAG ingestion and preview.
    """
    from fpdf import FPDF

    author = rows[0]["KONTO"] if rows else "AUTH-0000"
    agreement = rows[0]["AFTALE"] if rows else "AFT-0000-000"
    batch = rows[0]["AFREGNBATCH"] if rows else "0000"

    author_name = next((a[1] for a in AUTHORS if a[0] == author), "Unknown Author")
    title = random.choice(TITLES)

    base_date = _random_date(2020, 2026)
    period_start = base_date - timedelta(days=180)
    period_end = base_date

    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    # ── Header block ──
    pdf.set_font("Courier", "B", 11)
    pdf.cell(0, 6, "=" * 72, ln=True)
    pdf.set_font("Courier", "B", 12)
    pdf.cell(0, 7, "  SCHILLING  -  Royalty afregning", ln=True)
    pdf.set_font("Courier", "", 10)
    pdf.cell(0, 6, f"  Afregning nr: {batch}", ln=True)
    pdf.cell(0, 6, f"  Periode: {period_start.strftime('%d.%m.%y')}-{period_end.strftime('%d.%m.%y')}", ln=True)
    pdf.set_font("Courier", "B", 11)
    pdf.cell(0, 6, "=" * 72, ln=True)
    pdf.ln(3)

    # ── Meta fields ──
    pdf.set_font("Courier", "", 10)
    meta = [
        ("Titel:", title),
        ("Kontonr:", author),
        ("Forfatter:", author_name),
        ("Aftale:", agreement),
        ("Primo lager:", str(random.randint(100, 5000))),
        ("Ultimo lager:", str(random.randint(50, 4000))),
        ("Frieksemplarer:", str(random.randint(0, 50))),
        ("Makulatur:", str(random.randint(0, 200))),
    ]
    for label, value in meta:
        pdf.cell(0, 5, f"  {label:<18} {value}", ln=True)
    pdf.ln(3)

    # ── Table header ──
    pdf.set_font("Courier", "B", 9)
    pdf.cell(0, 5, "-" * 72, ln=True)
    header_line = f"  {'Salgskanal':<15} {'Prisgruppe':<12} {'Sats':<10} {'Antal':>8} {'Prisgrundlag':>14} {'Royalty':>12}"
    pdf.cell(0, 5, header_line, ln=True)
    pdf.cell(0, 5, "-" * 72, ln=True)

    # ── Table rows ──
    pdf.set_font("Courier", "", 9)
    total_royalty = 0.0
    for row in rows:
        if row["TRANSTYPE"] not in ("Salg", "Retur", "Frieksp", "Korrektion"):
            continue
        kanal = row["KANAL"]
        pris = row["PRISGRUPPE"]
        rate_str = row["STKAFREGNSATS"]
        try:
            rate_val = float(rate_str) if rate_str else 0.0
        except (ValueError, TypeError):
            rate_val = 0.0
        sats_display = f"{rate_val*100:.1f}%" if 0 < rate_val < 1 else str(rate_val)
        qty = row["ANTAL"]
        try:
            price = float(row["STKPRIS"])
            qty_val = int(float(qty))
            prisgrundlag = abs(qty_val) * price
        except (ValueError, TypeError):
            prisgrundlag = 0.0
        try:
            royalty = float(row["BELOEB"])
        except (ValueError, TypeError):
            royalty = 0.0
        total_royalty += royalty
        pdf.cell(0, 5, f"  {kanal:<15} {pris:<12} {sats_display:<10} {qty:>8} {prisgrundlag:>14,.2f} {royalty:>12,.2f}", ln=True)

    pdf.set_font("Courier", "B", 9)
    pdf.cell(0, 5, "-" * 72, ln=True)
    pdf.ln(3)

    # ── Summary ──
    fordeling_pct = round(random.uniform(0.5, 1.0), 2)
    fordeling_amount = round(total_royalty * fordeling_pct, 2)
    rest_garanti = round(random.uniform(-5000, 0), 2) if random.random() < 0.3 else 0.0
    afgift = round(random.uniform(-1000, 0), 2) if random.random() < 0.2 else 0.0
    til_udbetaling = round(fordeling_amount + rest_garanti - afgift, 2)

    pdf.set_font("Courier", "", 10)
    pdf.cell(0, 6, f"  Royalty fordeling: {fordeling_pct*100:.0f}% af {total_royalty:,.2f} -> {fordeling_amount:,.2f}", ln=True)
    pdf.cell(0, 6, f"  Rest global garanti: {rest_garanti:,.2f}", ln=True)
    pdf.cell(0, 6, f"  Afgift: {afgift:,.2f}", ln=True)
    pdf.set_font("Courier", "B", 10)
    pdf.cell(0, 6, f"  Til udbetaling: {til_udbetaling:,.2f}", ln=True)
    pdf.ln(2)
    pdf.set_font("Courier", "B", 11)
    pdf.cell(0, 6, "=" * 72, ln=True)

    pdf.output(str(path))


# ── Naming helpers ───────────────────────────────────────────────────────────

def _settlement_name(idx: int, scenario: str, fmt: str) -> str:
    """Generate a realistic filename."""
    patterns = [
        lambda: f"royalty_settlement_{idx:04d}",
        lambda: f"afregning_{_random_date().strftime('%Y%m%d')}_{idx:03d}",
        lambda: f"RS-{random.choice(AGREEMENTS).replace('AFT-', '')}_{idx:03d}",
        lambda: f"settlement_batch_{random.randint(1000,9999)}_{idx:03d}",
        lambda: f"royalty_{random.choice([a[0] for a in AUTHORS])}_{_random_date().strftime('%Y_%m')}",
        lambda: f"statement_{_random_date().strftime('%Y')}Q{random.randint(1,4)}_{idx:03d}",
        lambda: f"afregndata_{random.choice([a[0] for a in AUTHORS]).replace('AUTH-', 'A')}_{idx:03d}",
        lambda: f"export_schilling_{idx:04d}",
    ]
    base = random.choice(patterns)()
    return f"{base}.{fmt}"


# ── Main generation ──────────────────────────────────────────────────────────

def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    file_count = 0
    total = TARGET_CSV + TARGET_XLSX + TARGET_JSON + TARGET_PDF

    print(f"Generating {total} files in {OUTPUT_DIR} ...")

    # ---------- CSV files ----------
    for i in range(TARGET_CSV):
        scenario = _pick_scenario()
        num_rows = random.randint(10, 80)
        rows = _gen_file_rows(scenario, num_rows)
        delimiter = ";" if random.random() < 0.7 else ","
        name = _settlement_name(i, scenario, "csv")
        _write_csv(rows, OUTPUT_DIR / name, delimiter)
        file_count += 1
        if file_count % 50 == 0:
            print(f"  {file_count}/{total} files generated ...")

    # ---------- XLSX files ----------
    if HAS_OPENPYXL:
        for i in range(TARGET_XLSX):
            scenario = _pick_scenario()
            num_rows = random.randint(10, 80)
            rows = _gen_file_rows(scenario, num_rows)
            name = _settlement_name(TARGET_CSV + i, scenario, "xlsx")
            _write_xlsx(rows, OUTPUT_DIR / name)
            file_count += 1
            if file_count % 50 == 0:
                print(f"  {file_count}/{total} files generated ...")
    else:
        print("  ⚠ openpyxl not installed — skipping XLSX generation")
        # Generate extra CSV instead
        for i in range(TARGET_XLSX):
            scenario = _pick_scenario()
            num_rows = random.randint(10, 80)
            rows = _gen_file_rows(scenario, num_rows)
            name = _settlement_name(TARGET_CSV + i, scenario, "csv")
            _write_csv(rows, OUTPUT_DIR / name)
            file_count += 1
            if file_count % 50 == 0:
                print(f"  {file_count}/{total} files generated ...")

    # ---------- JSON files ----------
    for i in range(TARGET_JSON):
        scenario = _pick_scenario()
        num_rows = random.randint(10, 80)
        rows = _gen_file_rows(scenario, num_rows)
        name = _settlement_name(TARGET_CSV + TARGET_XLSX + i, scenario, "json")
        _write_json(rows, OUTPUT_DIR / name)
        file_count += 1
        if file_count % 50 == 0:
            print(f"  {file_count}/{total} files generated ...")

    # ---------- Pseudo-PDF files ----------
    for i in range(TARGET_PDF):
        scenario = _pick_scenario()
        num_rows = random.randint(15, 60)
        rows = _gen_file_rows(scenario, num_rows)
        name = _settlement_name(TARGET_CSV + TARGET_XLSX + TARGET_JSON + i, scenario, "pdf")
        _write_pseudo_pdf(rows, OUTPUT_DIR / name)
        file_count += 1
        if file_count % 50 == 0:
            print(f"  {file_count}/{total} files generated ...")

    print(f"\n✓ Done! Generated {file_count} files in {OUTPUT_DIR}")
    
    # Print stats
    exts = {}
    for f in OUTPUT_DIR.iterdir():
        ext = f.suffix.lower()
        exts[ext] = exts.get(ext, 0) + 1
    for ext, count in sorted(exts.items()):
        print(f"  {ext}: {count} files")


if __name__ == "__main__":
    main()
